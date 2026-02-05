# -*- coding: utf-8 -*-
"""
理赔查勘质量自动判定系统
根据《查勘质量考核基础表》规则，对案件数据进行自动化质量判定

使用方法：
    python claim_quality_checker.py 输入文件.xlsx
    
输出：
    在同目录下生成 "判定结果_输入文件.xlsx"
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ============== 险种分类映射 ==============
INSURANCE_CATEGORIES = {
    '雇主类': ['雇主责任'],
    '安责类': ['安全生产责任', '建筑施工行业安全生产责任', '建工团意'],
    '意外类': ['意外险'],
    '重疾类': ['重大疾病'],
    '财产类': ['财产一切', '财产综合', '财产基本', '机器损坏'],
    '工程类': ['建工一切', '路桥工程', '安工一切'],
    '公众类': ['公众责任'],
}

# ============== 核心要素关键词 ==============
CORE_ELEMENTS = {
    '雇主类': ['伤情', '部位', '医院', '岗位', '工种', '工伤', '劳动关系', '分包', '考勤'],
    '安责类': ['伤情', '部位', '医院', '岗位', '工种', '工伤', '劳动关系', '分包', '考勤'],
    '意外类': ['伤情', '部位', '医院', '职业'],
    '重疾类': ['诊断时间', '疾病名称', '疾病编码', '既往病史'],
    '财产类': ['标的类型', '原因类型', '维修方式', '财务账册', '残值'],
    '工程类': ['工程名称', '标的类型', '原因类型', '进度', '维修', '三者'],
    '公众类': ['归属', '标识', '监控', '三者', '医疗', '和解', '赔偿'],
    '其他类': ['损失类型', '后续处理'],
}

# ============== 必填项正则模式 ==============
MANDATORY_PATTERNS = {
    '查勘时间': r'查勘.{0,3}时间|于.{0,10}(查勘|勘查)|查勘.{0,5}日期',
    '查勘地点': r'查勘.{0,3}地点|在.{0,10}(查勘|勘查)|现场位于',
    '查勘方式': r'现场|视频|核实|走访|电话|远程|到场',
    '出险时间': r'出险.{0,3}时间|事故.{0,3}时间|发生.{0,5}时间|于.{0,10}(出险|发生)',
    '出险地点': r'出险.{0,3}地点|事故.{0,3}地点|发生.{0,5}地点|位于.{0,10}发生',
}

# ============== 报案延迟关键词 ==============
DELAY_KEYWORDS = ['延迟原因', '原因', '核实', '属实']

# ============== 机构排序规则 ==============
INSTITUTION_ORDER = ['合肥', '芜湖', '蚌埠', '淮南', '马鞍山', '淮北', '铜陵', '安庆', '黄山', 
                     '滁州', '阜阳', '宿州', '六安', '亳州', '池州', '宣城']


def classify_insurance(insurance_type: str) -> str:
    """险种分类"""
    if pd.isna(insurance_type):
        return '其他类'
    
    for category, keywords in INSURANCE_CATEGORIES.items():
        for keyword in keywords:
            if keyword in str(insurance_type):
                return category
    return '其他类'


def check_mandatory(survey_summary: str) -> tuple[bool, list]:
    """
    规则一：必填项目校验
    返回：(是否通过, 缺失项目列表)
    """
    if pd.isna(survey_summary):
        return False, list(MANDATORY_PATTERNS.keys())
    
    text = str(survey_summary)
    missing = []
    
    for item, pattern in MANDATORY_PATTERNS.items():
        if not re.search(pattern, text):
            missing.append(item)
    
    return len(missing) == 0, missing


def check_delay(accident_time, report_time, survey_summary: str) -> tuple[bool, bool, str]:
    """
    规则二：报案延迟风险识别
    返回：(是否触发延迟, 是否通过校验, 原因说明)
    """
    try:
        if pd.isna(accident_time) or pd.isna(report_time):
            return False, True, ''
        
        # 解析时间
        if isinstance(accident_time, str):
            accident_dt = pd.to_datetime(accident_time)
        else:
            accident_dt = pd.Timestamp(accident_time)
            
        if isinstance(report_time, str):
            report_dt = pd.to_datetime(report_time)
        else:
            report_dt = pd.Timestamp(report_time)
        
        # 计算延迟天数
        delay_days = (report_dt - accident_dt).days
        
        if delay_days <= 7:
            return False, True, ''
        
        # 触发延迟，检查查勘摘要
        if pd.isna(survey_summary):
            return True, False, f'报案延迟{delay_days}天，查勘摘要未填写延迟原因'
        
        text = str(survey_summary)
        for keyword in DELAY_KEYWORDS:
            if keyword in text:
                return True, True, ''
        
        return True, False, f'报案延迟{delay_days}天，查勘摘要未包含延迟原因说明'
        
    except Exception as e:
        return False, True, ''


def check_core_elements(category: str, survey_summary: str) -> tuple[bool, int, list]:
    """
    规则三：核心要素匹配
    返回：(是否通过, 匹配数量, 缺失要素列表)
    """
    if pd.isna(survey_summary):
        elements = CORE_ELEMENTS.get(category, CORE_ELEMENTS['其他类'])
        return False, 0, elements
    
    text = str(survey_summary)
    elements = CORE_ELEMENTS.get(category, CORE_ELEMENTS['其他类'])
    
    matched = []
    missing = []
    
    for element in elements:
        # 使用模糊匹配
        pattern = element.replace('类型', '.{0,2}类型?').replace('方式', '.{0,2}方式?')
        if re.search(pattern, text) or element in text:
            matched.append(element)
        else:
            missing.append(element)
    
    # 至少匹配3个核心要素
    min_required = min(3, len(elements))
    passed = len(matched) >= min_required
    
    return passed, len(matched), missing


def check_overlap(report_summary: str, survey_summary: str) -> tuple[bool, float]:
    """
    规则四：摘要重合率校验
    返回：(是否通过, 重合率)
    """
    if pd.isna(report_summary) or pd.isna(survey_summary):
        return True, 0.0
    
    report_text = str(report_summary).strip()
    survey_text = str(survey_summary).strip()
    
    if not report_text or not survey_text:
        return True, 0.0
    
    # 计算重合字符
    report_chars = set(report_text)
    survey_chars = set(survey_text)
    overlap_chars = report_chars & survey_chars
    
    # 重合率 = 重合字符数 / 较长文本的字符集大小
    max_len = max(len(report_chars), len(survey_chars))
    overlap_rate = len(overlap_chars) / max_len if max_len > 0 else 0
    
    # 重合率 < 80% 为合格（查勘摘要应有独立信息）
    return overlap_rate < 0.8, overlap_rate


def evaluate_case(row: pd.Series) -> tuple[str, str]:
    """
    综合评估单个案件
    返回：(判定结果, 不合格原因)
    """
    reasons = []
    
    # 获取必要字段
    insurance_type = row.get('险种', '')
    accident_time = row.get('出险时间', None)
    report_time = row.get('报案时间', None)
    report_summary = row.get('报案摘要', '')
    survey_summary = row.get('查勘摘要', '')
    
    # 险种分类
    category = classify_insurance(insurance_type)
    
    # 规则一：必填项校验
    mandatory_pass, mandatory_missing = check_mandatory(survey_summary)
    if not mandatory_pass:
        reasons.append(f"必填项缺失：{', '.join(mandatory_missing)}")
    
    # 规则二：报案延迟识别
    delay_triggered, delay_pass, delay_reason = check_delay(accident_time, report_time, survey_summary)
    if delay_triggered and not delay_pass:
        reasons.append(delay_reason)
    
    # 规则三：核心要素匹配
    core_pass, core_count, core_missing = check_core_elements(category, survey_summary)
    if not core_pass:
        min_required = min(3, len(CORE_ELEMENTS.get(category, CORE_ELEMENTS['其他类'])))
        reasons.append(f"核心要素不足（需{min_required}个，仅{core_count}个），建议补充：{', '.join(core_missing[:3])}")
    
    # 规则四：摘要重合率
    overlap_pass, overlap_rate = check_overlap(report_summary, survey_summary)
    if not overlap_pass:
        reasons.append(f"查勘摘要与报案摘要重合率过高（{overlap_rate:.1%}），缺乏独立调查信息")
    
    # 综合判定
    if reasons:
        return '不合格', '；'.join(reasons)
    else:
        return '合格', ''


def create_summary_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """创建机构汇总表"""
    # 按机构统计
    summary = df.groupby('机构').agg(
        总案件数=('判定结果', 'count'),
        合格案件数=('判定结果', lambda x: (x == '合格').sum())
    ).reset_index()
    
    summary['合格率'] = (summary['合格案件数'] / summary['总案件数'] * 100).round(1).astype(str) + '%'
    
    # 按预设顺序排序
    def sort_key(institution):
        try:
            return INSTITUTION_ORDER.index(institution)
        except ValueError:
            return len(INSTITUTION_ORDER)
    
    summary['排序'] = summary['机构'].apply(sort_key)
    summary = summary.sort_values('排序').drop('排序', axis=1)
    
    # 添加合计行
    total_cases = summary['总案件数'].sum()
    total_qualified = summary['合格案件数'].sum()
    total_rate = f"{total_qualified / total_cases * 100:.1f}%" if total_cases > 0 else "0%"
    
    total_row = pd.DataFrame({
        '机构': ['合计'],
        '总案件数': [total_cases],
        '合格案件数': [total_qualified],
        '合格率': [total_rate]
    })
    
    summary = pd.concat([summary, total_row], ignore_index=True)
    
    return summary


def style_excel(wb: Workbook):
    """美化 Excel 样式"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in wb.worksheets:
        # 设置表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置数据行样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """主程序入口"""
    # 获取输入文件
    if len(sys.argv) < 2:
        print("请将 Excel 文件拖拽到本程序上运行，或使用命令行：")
        print("  python claim_quality_checker.py 输入文件.xlsx")
        input("按回车键退出...")
        return
    
    input_file = Path(sys.argv[1])
    
    if not input_file.exists():
        print(f"错误：文件不存在 - {input_file}")
        input("按回车键退出...")
        return
    
    print(f"正在处理：{input_file.name}")
    
    try:
        # 读取 Excel
        df = pd.read_excel(input_file)
        print(f"读取到 {len(df)} 条记录")
        
        # 执行判定
        results = df.apply(evaluate_case, axis=1)
        df['判定结果'] = results.apply(lambda x: x[0])
        df['不合格原因'] = results.apply(lambda x: x[1])
        
        # 添加险种分类列（辅助分析）
        df['险种分类'] = df['险种'].apply(classify_insurance)
        
        # 创建汇总表
        summary_df = create_summary_sheet(df)
        
        # 输出结果
        output_file = input_file.parent / f"判定结果_{input_file.stem}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='机构汇总表', index=False)
            df.to_excel(writer, sheet_name='案件清单表', index=False)
        
        # 美化样式
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        style_excel(wb)
        wb.save(output_file)
        
        # 输出统计
        total = len(df)
        qualified = (df['判定结果'] == '合格').sum()
        print(f"\n判定完成！")
        print(f"  总案件数：{total}")
        print(f"  合格案件：{qualified}")
        print(f"  合格率：{qualified/total*100:.1f}%")
        print(f"\n结果已保存至：{output_file}")
        
    except Exception as e:
        print(f"处理出错：{e}")
        import traceback
        traceback.print_exc()
    
    input("\n按回车键退出...")


if __name__ == '__main__':
    main()
